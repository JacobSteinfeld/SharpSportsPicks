#!/usr/bin/env python3
"""
Generate SharpSportsPicks AI — Excel Dashboard from picks.csv
Run: python generate_excel.py
"""

import csv
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

DB_FILE  = "picks.csv"
OUT_FILE = "SharpSportsPicks_Dashboard.xlsx"

# ── Colors ────────────────────────────────────────────────────────────────────
C_WIN       = "C6EFCE"   # green fill
C_WIN_FONT  = "276221"
C_LOSS      = "FFC7CE"   # red fill
C_LOSS_FONT = "9C0006"
C_PUSH      = "FFEB9C"   # yellow fill
C_PUSH_FONT = "9C6500"
C_PENDING   = "DDEBF7"   # blue fill
C_PEND_FONT = "1F4E79"
C_HEADER    = "1F2D3D"   # dark navy header
C_HEADER_F  = "FFFFFF"
C_ACCENT    = "00B0F0"   # bright blue accent
C_SHEET_BG  = "F2F2F2"
C_SUMMARY_H = "2E4057"
C_SUMMARY_F = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_picks():
    with open(DB_FILE, newline="") as f:
        return list(csv.DictReader(f))


def build_picks_sheet(wb, picks):
    ws = wb.active
    ws.title = "All Picks"
    ws.sheet_view.showGridLines = False

    # ── Title bar ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "SharpSportsPicks AI — Pick Tracker"
    title_cell.fill    = fill(C_HEADER)
    title_cell.font    = Font(bold=True, color=C_HEADER_F, size=16, name="Calibri")
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:L2")
    sub = ws["A2"]
    sub.value = f"Generated {date.today().isoformat()}  |  All times in units"
    sub.fill  = fill("2E4057")
    sub.font  = Font(color="AAAAAA", size=10, name="Calibri", italic=True)
    sub.alignment = center()
    ws.row_dimensions[2].height = 18

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["#", "Date", "Sport", "Game", "Pick", "Team", "Bet Type",
               "Odds", "Units", "Result", "P&L", "Running P&L", "Notes"]
    col_widths = [5, 12, 8, 32, 20, 16, 10, 8, 7, 9, 9, 12, 45]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill      = fill(C_HEADER)
        cell.font      = Font(bold=True, color=C_HEADER_F, size=11, name="Calibri")
        cell.alignment = center()
        cell.border    = border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    running_pnl = 0.0
    for row_idx, p in enumerate(picks, start=4):
        result  = p["result"].lower()
        pnl_val = float(p["pnl"]) if p["pnl"] else 0.0
        if result in ("win", "loss", "push"):
            running_pnl += pnl_val

        # Row fill by result
        if result == "win":
            row_fill, row_font_color = C_WIN, C_WIN_FONT
        elif result == "loss":
            row_fill, row_font_color = C_LOSS, C_LOSS_FONT
        elif result == "push":
            row_fill, row_font_color = C_PUSH, C_PUSH_FONT
        else:
            row_fill, row_font_color = C_PENDING, C_PEND_FONT

        values = [
            int(p["id"]),
            p["date"],
            p["sport"],
            p["game"],
            p["pick"],
            p.get("team", ""),
            p["bet_type"].title(),
            p["odds"],
            float(p["units"]),
            p["result"].upper(),
            pnl_val if p["pnl"] else "",
            running_pnl if result in ("win","loss","push") else "",
            p["notes"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill(row_fill)
            cell.font      = Font(color=row_font_color, size=10, name="Calibri")
            cell.border    = border()
            cell.alignment = center() if col not in (4, 5, 12) else left()

            # P&L column — bold the number
            if col == 11 and val != "":
                cell.font = Font(bold=True, color=row_font_color, size=10, name="Calibri")
                cell.number_format = '+0.00;-0.00;0.00'
            if col == 12 and val != "":
                cell.number_format = '+0.00;-0.00;0.00'
                cell.font = Font(bold=True,
                                 color=C_WIN_FONT if running_pnl >= 0 else C_LOSS_FONT,
                                 size=10, name="Calibri")

        ws.row_dimensions[row_idx].height = 18

    # ── AutoFilter ────────────────────────────────────────────────────────────
    last_row = 3 + len(picks)
    ws.auto_filter.ref = f"A3:M{last_row}"

    # ── Freeze panes below header ──────────────────────────────────────────────
    ws.freeze_panes = "A4"

    return ws


def build_summary_sheet(wb, picks):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    settled = [p for p in picks if p["result"] in ("win", "loss", "push")]
    pending = [p for p in picks if p["result"] == "pending"]
    wins    = [p for p in settled if p["result"] == "win"]
    losses  = [p for p in settled if p["result"] == "loss"]
    total_pnl = sum(float(p["pnl"]) for p in settled if p["pnl"])
    win_rate  = (len(wins) / len(settled) * 100) if settled else 0

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "SharpSportsPicks AI — Summary"
    t.fill      = fill(C_HEADER)
    t.font      = Font(bold=True, color=C_HEADER_F, size=15, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    # Overall stats block
    stats = [
        ("Overall Record",  f"{len(wins)}-{len(losses)}"),
        ("Win Rate",        f"{win_rate:.1f}%"),
        ("Total P&L",       f"{total_pnl:+.2f}u"),
        ("Picks Logged",    str(len(picks))),
        ("Settled",         str(len(settled))),
        ("Pending",         str(len(pending))),
    ]

    ws.merge_cells("A2:E2")
    sec = ws["A2"]
    sec.value     = "OVERALL"
    sec.fill      = fill(C_SUMMARY_H)
    sec.font      = Font(bold=True, color=C_SUMMARY_F, size=11, name="Calibri")
    sec.alignment = center()

    for i, (label, val) in enumerate(stats, start=3):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=val)
        lc.font      = Font(bold=True, size=11, name="Calibri")
        lc.alignment = left()
        lc.border    = border()
        lc.fill      = fill("F2F2F2")
        vc.alignment = center()
        vc.border    = border()

        # Color P&L cell
        if label == "Total P&L":
            vc.fill = fill(C_WIN if total_pnl >= 0 else C_LOSS)
            vc.font = Font(bold=True,
                           color=C_WIN_FONT if total_pnl >= 0 else C_LOSS_FONT,
                           size=11, name="Calibri")
        elif label == "Win Rate":
            vc.fill = fill(C_WIN if win_rate >= 55 else C_LOSS if win_rate < 45 else C_PUSH)
        else:
            vc.fill = fill("FFFFFF")
        ws.row_dimensions[i].height = 20

    # Sport breakdown
    start_row = len(stats) + 4
    ws.merge_cells(f"A{start_row}:E{start_row}")
    sh = ws.cell(row=start_row, column=1, value="BY SPORT")
    sh.fill      = fill(C_SUMMARY_H)
    sh.font      = Font(bold=True, color=C_SUMMARY_F, size=11, name="Calibri")
    sh.alignment = center()

    sport_headers = ["Sport", "W", "L", "Win %", "P&L"]
    for col, h in enumerate(sport_headers, 1):
        c = ws.cell(row=start_row + 1, column=col, value=h)
        c.fill      = fill(C_HEADER)
        c.font      = Font(bold=True, color=C_HEADER_F, size=10, name="Calibri")
        c.alignment = center()
        c.border    = border()

    sports = {}
    for p in settled:
        s = p["sport"]
        if s not in sports:
            sports[s] = {"w": 0, "l": 0, "pnl": 0.0}
        if p["result"] == "win":
            sports[s]["w"] += 1
        elif p["result"] == "loss":
            sports[s]["l"] += 1
        sports[s]["pnl"] += float(p["pnl"]) if p["pnl"] else 0

    for i, (sport, d) in enumerate(sorted(sports.items()), start=start_row + 2):
        wr = d["w"] / (d["w"] + d["l"]) * 100 if (d["w"] + d["l"]) > 0 else 0
        row_data = [sport, d["w"], d["l"], f"{wr:.1f}%", f"{d['pnl']:+.2f}u"]
        pnl_fill = C_WIN if d["pnl"] >= 0 else C_LOSS
        pnl_font = C_WIN_FONT if d["pnl"] >= 0 else C_LOSS_FONT
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = center()
            c.border    = border()
            c.fill      = fill(pnl_fill if col == 5 else "FFFFFF")
            c.font      = Font(bold=(col == 5),
                               color=pnl_font if col == 5 else "000000",
                               size=10, name="Calibri")
        ws.row_dimensions[i].height = 18

    # Bet type breakdown
    bt_start = start_row + len(sports) + 4
    ws.merge_cells(f"A{bt_start}:E{bt_start}")
    bh = ws.cell(row=bt_start, column=1, value="BY BET TYPE")
    bh.fill      = fill(C_SUMMARY_H)
    bh.font      = Font(bold=True, color=C_SUMMARY_F, size=11, name="Calibri")
    bh.alignment = center()

    for col, h in enumerate(["Bet Type", "W", "L", "Win %", "P&L"], 1):
        c = ws.cell(row=bt_start + 1, column=col, value=h)
        c.fill      = fill(C_HEADER)
        c.font      = Font(bold=True, color=C_HEADER_F, size=10, name="Calibri")
        c.alignment = center()
        c.border    = border()

    bet_types = {}
    for p in settled:
        bt = p["bet_type"].title()
        if bt not in bet_types:
            bet_types[bt] = {"w": 0, "l": 0, "pnl": 0.0}
        if p["result"] == "win":
            bet_types[bt]["w"] += 1
        elif p["result"] == "loss":
            bet_types[bt]["l"] += 1
        bet_types[bt]["pnl"] += float(p["pnl"]) if p["pnl"] else 0

    for i, (bt, d) in enumerate(sorted(bet_types.items()), start=bt_start + 2):
        wr = d["w"] / (d["w"] + d["l"]) * 100 if (d["w"] + d["l"]) > 0 else 0
        row_data = [bt, d["w"], d["l"], f"{wr:.1f}%", f"{d['pnl']:+.2f}u"]
        pnl_fill = C_WIN if d["pnl"] >= 0 else C_LOSS
        pnl_font = C_WIN_FONT if d["pnl"] >= 0 else C_LOSS_FONT
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = center()
            c.border    = border()
            c.fill      = fill(pnl_fill if col == 5 else "FFFFFF")
            c.font      = Font(bold=(col == 5),
                               color=pnl_font if col == 5 else "000000",
                               size=10, name="Calibri")
        ws.row_dimensions[i].height = 18

    # Column widths
    for col, w in zip([1, 2, 3, 4, 5], [18, 8, 8, 10, 12]):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_teams_sheet(wb, picks):
    ws = wb.create_sheet("Team Tracker")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "SharpSportsPicks AI — If You Followed Every Team Play"
    t.fill      = fill(C_HEADER)
    t.font      = Font(bold=True, color=C_HEADER_F, size=14, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    headers    = ["Team", "Sport", "W", "L", "Win %", "P&L", "Picks"]
    col_widths = [22, 10, 6, 6, 8, 10, 7]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill      = fill(C_HEADER)
        c.font      = Font(bold=True, color=C_HEADER_F, size=11, name="Calibri")
        c.alignment = center()
        c.border    = border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    settled = [p for p in picks if p["result"] in ("win", "loss", "push")]
    teams   = {}
    for p in settled:
        t = p.get("team", "").strip()
        if not t:
            continue
        if t not in teams:
            teams[t] = {"w": 0, "l": 0, "pnl": 0.0, "picks": 0, "sport": p["sport"]}
        teams[t]["picks"] += 1
        teams[t]["pnl"]   += float(p["pnl"]) if p["pnl"] else 0
        if p["result"] == "win":
            teams[t]["w"] += 1
        elif p["result"] == "loss":
            teams[t]["l"] += 1

    sorted_teams = sorted(teams.items(), key=lambda x: x[1]["pnl"], reverse=True)

    for row_idx, (team, d) in enumerate(sorted_teams, start=3):
        total  = d["w"] + d["l"]
        wr     = d["w"] / total * 100 if total else 0
        pnl_f  = C_WIN if d["pnl"] >= 0 else C_LOSS
        pnl_fc = C_WIN_FONT if d["pnl"] >= 0 else C_LOSS_FONT
        row_bg = "F9F9F9" if row_idx % 2 == 0 else "FFFFFF"

        row_vals = [team, d["sport"], d["w"], d["l"], f"{wr:.0f}%", f"{d['pnl']:+.2f}u", d["picks"]]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = center()
            c.border    = border()
            if col == 6:
                c.fill = fill(pnl_f)
                c.font = Font(bold=True, color=pnl_fc, size=10, name="Calibri")
            else:
                c.fill = fill(row_bg)
                c.font = Font(size=10, name="Calibri")
        ws.row_dimensions[row_idx].height = 18

    if not sorted_teams:
        ws.merge_cells("A3:G3")
        msg = ws["A3"]
        msg.value     = "No team data yet — team field is now logged on every new pick."
        msg.alignment = center()
        msg.font      = Font(italic=True, color="888888", size=10, name="Calibri")

    ws.auto_filter.ref = f"A2:G{max(3, 2 + len(sorted_teams))}"
    ws.freeze_panes    = "A3"


def build_pitchers_sheet(wb, picks):
    ws = wb.create_sheet("Pitcher Tracker")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = "SharpSportsPicks AI — MLB Pitcher Tracker"
    t.fill      = fill(C_HEADER)
    t.font      = Font(bold=True, color=C_HEADER_F, size=14, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # Section headers
    ws.merge_cells("A2:E2")
    a = ws["A2"]
    a.value     = "ALL STARTS"
    a.fill      = fill(C_SUMMARY_H)
    a.font      = Font(bold=True, color=C_SUMMARY_F, size=10, name="Calibri")
    a.alignment = center()

    ws.merge_cells("F2:I2")
    m = ws["F2"]
    m.value     = "★  MARQUEE STARTS ONLY"
    m.fill      = fill("8B4513")
    m.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    m.alignment = center()

    headers    = ["Pitcher", "W", "L", "Win%", "P&L", "W", "L", "Win%", "P&L"]
    col_widths = [24, 6, 6, 8, 10, 6, 6, 8, 10]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill      = fill(C_HEADER)
        c.font      = Font(bold=True, color=C_HEADER_F, size=10, name="Calibri")
        c.alignment = center()
        c.border    = border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 18

    mlb = [p for p in picks if p.get("sport") == "MLB" and p.get("result") in ("win","loss","push")]
    pitchers = {}
    for p in mlb:
        name = p.get("pitcher", "").strip()
        if not name:
            continue
        if name not in pitchers:
            pitchers[name] = {"w":0,"l":0,"pnl":0.0,"mq_w":0,"mq_l":0,"mq_pnl":0.0}
        d   = pitchers[name]
        pnl = float(p["pnl"]) if p["pnl"] else 0.0
        d["pnl"] += pnl
        if p["result"] == "win":  d["w"] += 1
        elif p["result"] == "loss": d["l"] += 1
        if p.get("marquee") == "yes":
            d["mq_pnl"] += pnl
            if p["result"] == "win":   d["mq_w"] += 1
            elif p["result"] == "loss": d["mq_l"] += 1

    sorted_p = sorted(pitchers.items(), key=lambda x: x[1]["pnl"], reverse=True)

    for row_idx, (name, d) in enumerate(sorted_p, start=4):
        total  = d["w"] + d["l"]
        wr     = d["w"] / total * 100 if total else 0
        mq_tot = d["mq_w"] + d["mq_l"]
        mq_wr  = d["mq_w"] / mq_tot * 100 if mq_tot else 0
        pnl_f  = C_WIN if d["pnl"] >= 0 else C_LOSS
        pnl_fc = C_WIN_FONT if d["pnl"] >= 0 else C_LOSS_FONT
        mq_f   = C_WIN if d["mq_pnl"] >= 0 else C_LOSS
        mq_fc  = C_WIN_FONT if d["mq_pnl"] >= 0 else C_LOSS_FONT
        bg     = "F9F9F9" if row_idx % 2 == 0 else "FFFFFF"

        row_vals = [
            name, d["w"], d["l"], f"{wr:.0f}%", f"{d['pnl']:+.2f}u",
            d["mq_w"] if mq_tot else "—",
            d["mq_l"] if mq_tot else "—",
            f"{mq_wr:.0f}%" if mq_tot else "—",
            f"{d['mq_pnl']:+.2f}u" if mq_tot else "—",
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = center()
            c.border    = border()
            if col == 5:
                c.fill = fill(pnl_f)
                c.font = Font(bold=True, color=pnl_fc, size=10, name="Calibri")
            elif col == 9:
                c.fill = fill(mq_f) if mq_tot else fill(bg)
                c.font = Font(bold=True, color=mq_fc if mq_tot else "AAAAAA", size=10, name="Calibri")
            elif col >= 6:
                c.fill = fill("FFF8E7")  # warm tint for marquee cols
                c.font = Font(size=10, name="Calibri")
            else:
                c.fill = fill(bg)
                c.font = Font(size=10, name="Calibri")
        ws.row_dimensions[row_idx].height = 18

    if not sorted_p:
        ws.merge_cells("A4:I4")
        msg = ws["A4"]
        msg.value     = "No MLB pitcher data yet — pitcher + marquee fields log automatically on MLB picks."
        msg.alignment = center()
        msg.font      = Font(italic=True, color="888888", size=10, name="Calibri")

    ws.freeze_panes = "A4"


def build_batters_sheet(wb, picks):
    ws = wb.create_sheet("Batter Props")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value     = "SharpSportsPicks AI — MLB Batter Prop Tracker"
    t.fill      = fill(C_HEADER)
    t.font      = Font(bold=True, color=C_HEADER_F, size=14, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    headers    = ["Batter", "Prop Type", "Line", "Opp Pitcher", "Batter", "Pitcher", "Game", "Result", "P&L", "Notes"]
    subheaders = ["",       "",          "",     "",             "Hand",   "Hand",    "",     "",       "",    ""]
    col_widths = [22, 14, 16, 22, 8, 8, 28, 9, 10, 35]

    for col, (h, sh, w) in enumerate(zip(headers, subheaders, col_widths), 1):
        c = ws.cell(row=2, column=col, value=h + (f"\n{sh}" if sh else ""))
        c.fill      = fill(C_HEADER)
        c.font      = Font(bold=True, color=C_HEADER_F, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 28

    props = [p for p in picks if p.get("batter","").strip()]

    for row_idx, p in enumerate(props, start=3):
        result = p["result"].lower()
        pnl    = float(p["pnl"]) if p["pnl"] else 0.0
        if result == "win":
            rf, rfc = C_WIN, C_WIN_FONT
        elif result == "loss":
            rf, rfc = C_LOSS, C_LOSS_FONT
        elif result == "push":
            rf, rfc = C_PUSH, C_PUSH_FONT
        else:
            rf, rfc = C_PENDING, C_PEND_FONT

        row_vals = [
            p.get("batter",""),
            p.get("prop_type","").title(),
            p.get("prop_line",""),
            p.get("pitcher",""),
            p.get("batter_hand",""),
            p.get("pitcher_hand",""),
            p.get("game",""),
            result.upper(),
            pnl if p["pnl"] else "",
            p.get("notes",""),
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = fill(rf)
            c.font      = Font(color=rfc, size=10, name="Calibri",
                               bold=(col == 9))
            c.alignment = left() if col in (1, 3, 7, 10) else center()
            c.border    = border()
            if col == 9 and val != "":
                c.number_format = '+0.00;-0.00;0.00'
        ws.row_dimensions[row_idx].height = 18

    if not props:
        ws.merge_cells("A3:J3")
        msg = ws["A3"]
        msg.value     = "No batter prop data yet — select 'Batter prop? y' when adding an MLB pick."
        msg.alignment = center()
        msg.font      = Font(italic=True, color="888888", size=10, name="Calibri")

    last = max(3, 2 + len(props))
    ws.auto_filter.ref = f"A2:J{last}"
    ws.freeze_panes    = "A3"

    # ── Prop type summary table ───────────────────────────────────────────────
    sum_start = last + 3
    ws.merge_cells(f"A{sum_start}:E{sum_start}")
    sh = ws.cell(row=sum_start, column=1, value="PROP TYPE SUMMARY")
    sh.fill = fill(C_SUMMARY_H); sh.font = Font(bold=True, color=C_SUMMARY_F, size=10, name="Calibri")
    sh.alignment = center()

    for col, h in enumerate(["Prop Type","W","L","Win%","P&L"], 1):
        c = ws.cell(row=sum_start+1, column=col, value=h)
        c.fill = fill(C_HEADER); c.font = Font(bold=True, color=C_HEADER_F, size=10, name="Calibri")
        c.alignment = center(); c.border = border()

    settled_props = [p for p in props if p["result"] in ("win","loss","push")]
    prop_types = {}
    for p in settled_props:
        pt = p.get("prop_type","").strip().title() or "Other"
        prop_types.setdefault(pt, {"w":0,"l":0,"pnl":0.0})
        prop_types[pt]["pnl"] += float(p["pnl"]) if p["pnl"] else 0
        if p["result"] == "win":   prop_types[pt]["w"] += 1
        elif p["result"] == "loss": prop_types[pt]["l"] += 1

    for i, (pt, d) in enumerate(sorted(prop_types.items(), key=lambda x: x[1]["pnl"], reverse=True), start=sum_start+2):
        total = d["w"] + d["l"]
        wr    = d["w"] / total * 100 if total else 0
        pf    = C_WIN if d["pnl"] >= 0 else C_LOSS
        pfc   = C_WIN_FONT if d["pnl"] >= 0 else C_LOSS_FONT
        for col, val in enumerate([pt, d["w"], d["l"], f"{wr:.0f}%", f"{d['pnl']:+.2f}u"], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = center(); c.border = border()
            c.fill = fill(pf if col==5 else "FFFFFF")
            c.font = Font(bold=(col==5), color=pfc if col==5 else "000000", size=10, name="Calibri")
        ws.row_dimensions[i].height = 18


def main():
    picks = load_picks()
    wb    = Workbook()

    build_picks_sheet(wb, picks)
    build_summary_sheet(wb, picks)
    build_teams_sheet(wb, picks)
    build_pitchers_sheet(wb, picks)
    build_batters_sheet(wb, picks)

    wb.save(OUT_FILE)
    print(f"✓ Dashboard saved → {OUT_FILE}")
    print(f"  {len(picks)} picks | Sheets: All Picks, Summary, Team Tracker, Pitcher Tracker, Batter Props")


if __name__ == "__main__":
    main()
